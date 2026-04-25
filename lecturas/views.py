from django.shortcuts import render
from .models import Caragral
from django.http import HttpRequest
from .forms import Formulario # este comando importa el formulario
#  que esta creado en la carpeta forms
from django.template.loader import get_template
from django.http import HttpResponse
from openpyxl import workbook

# Create your views here.



# Create your views here.
def Formulario2(self):
    plantilla= get_template('lecturas/formulario.html')
    return HttpResponse(plantilla.render())


# esta vista muestra todos los abonados
def listar_lecturas(request):
    
    lista_caragral=Caragral.objects.all()
    return render(request, "lecturas/lecturas.html",{"lista_caragral": lista_caragral})

#esta vista mostrara solo los abonados que faltan de ingresar
def listar_pendientes(request):

    lista_caragral=Caragral.objects.filter(Lectura_actual__exact=0)
    return render(request,'lecturas/lecturas.html', {'lista_caragral': lista_caragral})

# vista para modificar el abonado
#def edit(request,id_abonado):

#    abonado=Caragral.objects.filter(id=id_abonado).first()
#    form=Formulario(isinstance=abonado)
#    return render(request,"",{"form":form, 'abonado':abonado})

# esto es para poder trabajar con formularios
class FormularioView(HttpRequest):
#    @staticmethod 
#    se vana crear dos metodos     
    def index(request):
        abonado= Formulario()
        return render(request,"lecturas/abonadoindex.html", {"form":abonado})
    
#   se va a crear el segundo metodo    
    def procesar_formulario(request):
        abonado=Formulario(request.POST)
        if abonado.is_valid():
            abonado.save()  # guarda el abonado en la base de datos
            abonado=Formulario() # limpia el contenido del formulario

        return render(request,"lecturas/abonadoindex.html", {"form":abonado, "mensaje": 'O.K.'})

    def edit(request,id_abonado):

        abonado=Caragral.objects.filter(id=id_abonado).first()
        form=Formulario(instance=abonado)
        return render(request,'lecturas/abonadoedit.html', {"form":form,'abonado':abonado})
    

#  se va a crear metodo o vista para actualizar el abonado
    def actualizar_abonado(request,id_abonado):

        abonado=Caragral.objects.get(pk=id_abonado)   
        form=Formulario(request.POST, instance=abonado)
        if form.is_valid():
            form.save()
        lista_caragral= Caragral.objects.all()
        return render(request, "lecturas/lecturas.html", {"lista_caragral": lista_caragral})

# voy a crear la vista para exportar a Excel las lectura
    def  exportar_excel(request):
        def get(self,request,*args,**kwargs):
            abonados=Caragral.objects.all()
            wb=workbook()
            ws=wb.active

            ws['A1'] = 'ABONADOID'
            ws['B1'] = 'NOMBRE'
            ws['C1'] = 'LECTURA ANTERIOR'
            ws['D1'] = 'LECTURA ACTUAL'
            ws['E1'] = 'OBSERVACIONES'

            cont= 2

            for abonado in abonados:
                ws.cell(row=cont, column=1).value = abonado.Abonadoid
                ws.cell(row=cont, column=2).value = abonado.Nombre
                ws.cell(row=cont, column=3).value = abonado.Lectura_anterior
                ws.cell(row=cont, column=4).value = abonado.Lectura_actual
                ws.cell(row=cont, column=1).value = abonado.Observacion
                cont+=1

            nombre_archivo="Exportar_hacia_excel.xlsx"
            response=HttpResponse(content_type="application/ms-excel")
            content="attachment; filename={0}".format(nombre_archivo)
            response['Content-Disposition']=content
            wb.save(response)
            return response
